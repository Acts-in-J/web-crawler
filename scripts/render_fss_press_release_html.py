from __future__ import annotations

import argparse
import html
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_workbook_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        record: dict[str, str] = {}
        for header, value in zip(headers, raw_row):
            record[header] = "" if value is None else str(value).strip()
        if any(record.values()):
            records.append(record)
    return records


def read_attachments(path: Path | None) -> dict[str, list[dict[str, str]]]:
    if path is None or not path.exists():
        return {}

    data = json.loads(path.read_text(encoding="utf-8"))
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for item in data.get("attachments", []):
        post_number = str(item.get("post_number", "")).strip()
        if not post_number:
            continue
        grouped[post_number].append(
            {
                "name": str(item.get("source_file_name", "")).strip(),
                "url": str(item.get("download_url", "")).strip(),
                "saved_path": str(item.get("saved_path", "")).strip(),
            }
        )
    return grouped


def read_validation(path: Path | None) -> dict[str, Any]:
    if path is None or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def paragraphs(text: str) -> list[str]:
    normalized = normalize_text(text)
    if not normalized:
        return []
    parts = re.split(r"(?=\s*[ㅁㅇ※]\s)", normalized)
    return [part.strip() for part in parts if part.strip()]


def local_pdf_href(saved_path: str, output_dir: Path) -> str:
    if not saved_path:
        return ""
    saved = Path(saved_path)
    try:
        return saved.resolve().relative_to(output_dir.resolve()).as_posix()
    except ValueError:
        return saved.as_uri()


def render_html(
    rows: list[dict[str, str]],
    attachments_by_number: dict[str, list[dict[str, str]]],
    validation: dict[str, Any],
    xlsx_path: Path,
    output_path: Path,
) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_url = validation.get("source_url", "")
    row_count = len(rows)
    pdf_count = sum(len(items) for items in attachments_by_number.values())
    departments = sorted({row.get("담당부서", "") for row in rows if row.get("담당부서", "")})
    dates = [row.get("등록일", "") for row in rows if row.get("등록일", "")]
    date_range = f"{min(dates)} ~ {max(dates)}" if dates else "-"

    cards = []
    for index, row in enumerate(rows, start=1):
        number = row.get("번호", "")
        title = row.get("제목", "")
        department = row.get("담당부서", "")
        date = row.get("등록일", "")
        body = row.get("본문 텍스트", "")
        source_link = row.get("본문 링크 url", "")
        attachments = attachments_by_number.get(number, [])
        search_text = " ".join([number, title, department, date, body])

        paragraph_html = "\n".join(
            f"            <p>{html.escape(part)}</p>" for part in paragraphs(body)
        )
        if not paragraph_html:
            paragraph_html = "            <p class=\"empty\">본문 텍스트 없음</p>"

        attachment_html = ""
        if attachments:
            attachment_items = []
            for attachment in attachments:
                local_href = local_pdf_href(attachment["saved_path"], output_path.parent)
                label = attachment["name"] or Path(attachment["saved_path"]).name or "첨부 PDF"
                links = []
                if local_href:
                    links.append(f"<a href=\"{html.escape(local_href)}\">로컬 PDF</a>")
                if attachment["url"]:
                    links.append(f"<a href=\"{html.escape(attachment['url'])}\" target=\"_blank\" rel=\"noopener\">원문 다운로드</a>")
                attachment_items.append(
                    "              <li>"
                    f"<span>{html.escape(label)}</span>"
                    f"<small>{' · '.join(links)}</small>"
                    "</li>"
                )
            attachment_html = (
                "          <section class=\"attachments\" aria-label=\"첨부 PDF\">\n"
                "            <h3>첨부 PDF</h3>\n"
                "            <ul>\n"
                + "\n".join(attachment_items)
                + "\n            </ul>\n"
                "          </section>"
            )

        source_link_html = (
            f"<a class=\"source-link\" href=\"{html.escape(source_link)}\" target=\"_blank\" rel=\"noopener\">원문 보기</a>"
            if source_link
            else ""
        )

        cards.append(
            f"""        <article class="release-card" data-search="{html.escape(search_text.lower())}" data-department="{html.escape(department)}">
          <header>
            <div class="rank">{index:02d}</div>
            <div>
              <p class="meta">번호 {html.escape(number)} · {html.escape(date)} · {html.escape(department)}</p>
              <h2>{html.escape(title)}</h2>
            </div>
          </header>
          <div class="body">
{paragraph_html}
          </div>
          <footer>
            {source_link_html}
          </footer>
{attachment_html}
        </article>"""
        )

    department_options = "\n".join(
        f"          <option value=\"{html.escape(department)}\">{html.escape(department)}</option>"
        for department in departments
    )

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>금융감독원 보도자료 50건</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #17202a;
      --muted: #5a6675;
      --line: #d8dde5;
      --panel: #ffffff;
      --wash: #f5f7fa;
      --accent: #0b5cab;
      --accent-ink: #073963;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Malgun Gothic", Arial, sans-serif;
      color: var(--ink);
      background: var(--wash);
      line-height: 1.58;
    }}
    a {{ color: var(--accent); text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .page {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 32px 20px 56px;
    }}
    .hero {{
      padding: 28px 0 20px;
      border-bottom: 1px solid var(--line);
    }}
    .eyebrow {{
      margin: 0 0 8px;
      color: var(--accent-ink);
      font-size: 13px;
      font-weight: 700;
      letter-spacing: 0;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(30px, 4vw, 46px);
      line-height: 1.12;
      letter-spacing: 0;
    }}
    .hero p {{
      max-width: 820px;
      margin: 14px 0 0;
      color: var(--muted);
      font-size: 15px;
    }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      margin: 22px 0;
    }}
    .summary div {{
      min-height: 88px;
      padding: 16px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }}
    .summary dt {{
      margin: 0;
      color: var(--muted);
      font-size: 12px;
    }}
    .summary dd {{
      margin: 6px 0 0;
      font-size: 22px;
      font-weight: 750;
    }}
    .controls {{
      position: sticky;
      top: 0;
      z-index: 5;
      display: grid;
      grid-template-columns: minmax(220px, 1fr) minmax(180px, 280px);
      gap: 12px;
      padding: 14px 0;
      background: var(--wash);
      border-bottom: 1px solid var(--line);
    }}
    input, select {{
      width: 100%;
      height: 42px;
      padding: 0 12px;
      border: 1px solid #b9c2cf;
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      font: inherit;
    }}
    .count {{
      margin: 14px 0;
      color: var(--muted);
      font-size: 13px;
    }}
    .release-list {{
      display: grid;
      gap: 14px;
    }}
    .release-card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 18px;
    }}
    .release-card header {{
      display: grid;
      grid-template-columns: 48px 1fr;
      gap: 14px;
      align-items: start;
    }}
    .rank {{
      display: grid;
      place-items: center;
      width: 42px;
      height: 42px;
      border: 1px solid #c3d5e8;
      border-radius: 50%;
      color: var(--accent-ink);
      font-weight: 800;
      font-size: 14px;
    }}
    .meta {{
      margin: 0 0 4px;
      color: var(--muted);
      font-size: 13px;
    }}
    h2 {{
      margin: 0;
      font-size: 20px;
      line-height: 1.35;
      letter-spacing: 0;
    }}
    .body {{
      margin-top: 14px;
      padding-left: 62px;
    }}
    .body p {{
      margin: 8px 0;
      word-break: keep-all;
      overflow-wrap: anywhere;
    }}
    .empty {{ color: var(--muted); }}
    footer {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 14px;
      padding-left: 62px;
    }}
    .source-link {{
      display: inline-flex;
      align-items: center;
      min-height: 34px;
      padding: 6px 10px;
      border: 1px solid #b7cee6;
      border-radius: 6px;
      background: #f3f8fd;
      font-weight: 700;
      font-size: 13px;
    }}
    .attachments {{
      margin-top: 14px;
      margin-left: 62px;
      padding: 12px;
      border: 1px solid #e1e6ee;
      border-radius: 8px;
      background: #fbfcfe;
    }}
    .attachments h3 {{
      margin: 0 0 8px;
      font-size: 14px;
    }}
    .attachments ul {{
      display: grid;
      gap: 8px;
      margin: 0;
      padding-left: 18px;
    }}
    .attachments li span {{
      display: block;
      word-break: keep-all;
      overflow-wrap: anywhere;
    }}
    .attachments small {{
      display: block;
      color: var(--muted);
      font-size: 12px;
    }}
    .hidden {{ display: none; }}
    .source-note {{
      margin-top: 22px;
      color: var(--muted);
      font-size: 12px;
      word-break: break-all;
    }}
    @media (max-width: 760px) {{
      .page {{ padding: 22px 14px 40px; }}
      .summary {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .controls {{ grid-template-columns: 1fr; }}
      .release-card header {{ grid-template-columns: 1fr; }}
      .rank {{ width: 36px; height: 36px; }}
      .body, footer, .attachments {{ padding-left: 0; margin-left: 0; }}
      h2 {{ font-size: 18px; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <p class="eyebrow">금융감독원 보도자료 수집 결과</p>
      <h1>최근 보도자료 50건 구조화 HTML</h1>
      <p>엑셀 파일의 행 순서를 유지하고, 각 보도자료의 본문 요약과 원문 링크, 다운로드된 PDF 첨부를 함께 묶었습니다.</p>
    </section>

    <dl class="summary">
      <div><dt>보도자료</dt><dd>{row_count}건</dd></div>
      <div><dt>첨부 PDF</dt><dd>{pdf_count}개</dd></div>
      <div><dt>등록일 범위</dt><dd>{html.escape(date_range)}</dd></div>
      <div><dt>생성 시각</dt><dd>{html.escape(generated_at)}</dd></div>
    </dl>

    <section class="controls" aria-label="필터">
      <input id="search" type="search" placeholder="제목, 부서, 본문 검색">
      <select id="department">
        <option value="">전체 담당부서</option>
{department_options}
      </select>
    </section>
    <p class="count"><span id="visible-count">{row_count}</span> / {row_count}건 표시</p>

    <section class="release-list" id="release-list">
{chr(10).join(cards)}
    </section>

    <p class="source-note">
      원본 엑셀: {html.escape(str(xlsx_path.resolve()))}<br>
      수집 URL: {html.escape(str(source_url))}
    </p>
  </main>

  <script>
    const searchInput = document.getElementById('search');
    const departmentSelect = document.getElementById('department');
    const cards = [...document.querySelectorAll('.release-card')];
    const visibleCount = document.getElementById('visible-count');

    function applyFilters() {{
      const query = searchInput.value.trim().toLowerCase();
      const department = departmentSelect.value;
      let shown = 0;

      for (const card of cards) {{
        const matchesQuery = !query || card.dataset.search.includes(query);
        const matchesDepartment = !department || card.dataset.department === department;
        const visible = matchesQuery && matchesDepartment;
        card.classList.toggle('hidden', !visible);
        if (visible) shown += 1;
      }}

      visibleCount.textContent = shown;
    }}

    searchInput.addEventListener('input', applyFilters);
    departmentSelect.addEventListener('change', applyFilters);
  </script>
</body>
</html>
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render FSS press-release crawl output as structured HTML.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--raw-data", type=Path)
    parser.add_argument("--validation", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = read_workbook_rows(args.xlsx)
    attachments = read_attachments(args.raw_data)
    validation = read_validation(args.validation)
    html_text = render_html(rows, attachments, validation, args.xlsx, args.output)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(html_text, encoding="utf-8", newline="\n")
    print(f"wrote {args.output} ({len(rows)} rows, {sum(len(v) for v in attachments.values())} pdfs)")


if __name__ == "__main__":
    main()
